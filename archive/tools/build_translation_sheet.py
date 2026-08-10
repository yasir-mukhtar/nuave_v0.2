#!/usr/bin/env python3
"""
Build an EN-ID translation review sheet from AppEn.tsx / AppId.tsx.

Columns:
  A: English
  B: Indonesian
  C: (empty) — user's revision goes here

Extraction strategy (robust against JSX quirks):
  1. Strip JSX comments.
  2. Isolate the JSX body: everything between the first `return (` and the
     last `);` — this excludes TypeScript code (imports, export function).
  3. Remove JSX expression braces {…} iteratively (handles nesting like
     {cost && (<CostBadge … />)}).
  4. Tokenize into tags and text; coalesce text runs into paragraph units:
     text separated only by inline tags (<strong>, <em>, <br>, <span>) stays
     in the same unit; any other tag closes the unit. This makes EN and ID
     unit counts match even where the translation moved an <em>/<strong>.
  5. Screen-visible string props (label, title, cost, tier, outOfLabel,
     scoreLabel) become their own units at their document positions.
  6. Zip EN and ID unit lists 1:1. Assert equal counts and matching kinds —
     if they differ, report the first divergence so the JSX can be fixed.
"""
import re
import html as html_mod
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path("/Users/hy4-mac-006/nuave_v0.2/report-prototype/src")
OUT = Path("/Users/hy4-mac-006/nuave_v0.2/report-en-id-review.xlsx")

INLINE_TAG = re.compile(r"^</?(?:strong|em|b|i|br|code)\b", re.I)
TAG = re.compile(r"(<[^>]+>)")
PROP = re.compile(
    r"\b(label|title|cost|tier|outOfLabel|scoreLabel)\s*=\s*\"([^\"]*)\""
)


def jsx_body(raw: str) -> str:
    """Extract the JSX returned by the component, dropping TS code."""
    raw = re.sub(r"\{/\*.*?\*/\}", "", raw, flags=re.S)
    m = re.search(r"return\s*\(", raw)
    if not m:
        raise ValueError("no `return (` found")
    start = m.end()
    end = raw.rfind(");")
    if end <= start:
        raise ValueError("no closing `);` found")
    body = raw[start:end]
    # remove JSX expression braces iteratively (handles nesting)
    prev = None
    while prev != body:
        prev = body
        body = re.sub(r"\{[^{}]*\}", "", body)
    return body


def extract_units(path: Path):
    """Return ordered (kind, value) paragraph units from a JSX file."""
    raw = path.read_text(encoding="utf-8")
    body = jsx_body(raw)

    units = []  # (position, kind, value)
    current = []  # text fragments of the open unit
    current_pos = None

    def flush():
        nonlocal current, current_pos
        if current:
            text = html_mod.unescape(" ".join(t for _, t in current))
            text = re.sub(r"\s+", " ", text).strip()
            if text:
                units.append((current_pos, "text", text))
        current = []
        current_pos = None

    pos = 0
    for tok in TAG.split(body):
        if not tok:
            continue
        if tok.startswith("<"):
            if INLINE_TAG.match(tok):
                pass  # inline tag: keep current unit open
            else:
                flush()  # block-level tag closes the unit
            pos += len(tok)
        else:
            # text fragment (inside the JSX body, so never TS code)
            t = re.sub(r"\s+", " ", tok).strip()
            if t:
                if current_pos is None:
                    current_pos = pos
                current.append((pos, t))
            pos += len(tok)

    flush()

    # screen-visible props become their own units at their positions
    for m in PROP.finditer(body):
        value = html_mod.unescape(m.group(2)).strip()
        if value:
            units.append((m.start(0), "prop:" + m.group(1), value))

    units.sort(key=lambda x: x[0])
    return [(k, v) for _, k, v in units]


en = extract_units(SRC / "AppEn.tsx")
id_ = extract_units(SRC / "AppId.tsx")
print(f"EN units: {len(en)}  ID units: {len(id_)}")

if len(en) != len(id_):
    print("!! LENGTH MISMATCH — first divergence:")
    for i in range(max(len(en), len(id_))):
        a = en[i] if i < len(en) else ("<missing>", "<missing>")
        b = id_[i] if i < len(id_) else ("<missing>", "<missing>")
        if a[0] != b[0]:
            print(f"  [{i}] EN: {a!r}")
            print(f"  [{i}] ID: {b!r}")
            break
    raise SystemExit(1)

pairs = list(zip(en, id_))
# Identical values are fine for proper nouns, numbers, loanwords
# ("Masryef", "56", "vs Amanie", "Status", client-name lists). Flag only
# long identical rows that contain English function words — those indicate
# a sentence that was never translated.
FUNCTION_WORDS = re.compile(
    r"\b(the|and|for|your|this|that|are|is|of|to|in|with|you|we|what|how|who|it|a|an)\b",
    re.I,
)
identical_long = [
    i for i, (a, b) in enumerate(pairs)
    if a[1] == b[1] and len(a[1]) > 30 and FUNCTION_WORDS.search(a[1])
]
if identical_long:
    print(f"!! Untranslated-looking EN/ID rows at {len(identical_long)} (should be 0):")
    for i in identical_long:
        print(f"  [{i}] {en[i][1]!r}")
    raise SystemExit(1)

print(f"Aligned pairs: {len(pairs)}")

# ── Build workbook ──
wb = Workbook()
ws = wb.active
ws.title = "Translation Review"

header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
body_font = Font(name="Arial", size=10)
section_font = Font(name="Arial", size=10, bold=True, color="1F4E78")
section_fill = PatternFill("solid", fgColor="DCE6F1")
rev_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")

headers = ["English", "Indonesian", "My revision"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = border

section_keys = [
    "Executive Summary", "Ringkasan Eksekutif",
    "Section 1", "Bagian 1",
    "Section 2", "Bagian 2",
    "Section 3", "Bagian 3",
    "Section 4", "Bagian 4",
    "Section 5", "Bagian 5",
    "Section 6", "Bagian 6",
    "Section 7", "Bagian 7",
    "Section 8", "Bagian 8",
    "Section 9", "Bagian 9",
    "Appendix A", "Lampiran A",
    "Appendix B", "Lampiran B",
]

r = 2
for (en_kind, en_text), (id_kind, id_text) in pairs:
    is_section = any(
        (en_text.strip() == k or id_text.strip() == k) for k in section_keys
    )
    if is_section:
        ws.cell(row=r, column=1, value=en_text).font = section_font
        ws.cell(row=r, column=2, value=id_text).font = section_font
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = section_fill
            cell.border = border
            cell.alignment = wrap_top
    else:
        c1 = ws.cell(row=r, column=1, value=en_text)
        c2 = ws.cell(row=r, column=2, value=id_text)
        c3 = ws.cell(row=r, column=3, value=None)
        for cell in (c1, c2, c3):
            cell.font = body_font
            cell.alignment = wrap_top
            cell.border = border
        c3.fill = rev_fill
    r += 1

ws.column_dimensions["A"].width = 62
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 40
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{r-1}"

wb.save(OUT)
print(f"Saved: {OUT}  (rows: {r-1})")
